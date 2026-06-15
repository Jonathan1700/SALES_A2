"""Mixins reutilizables para las vistas del módulo billing."""
from django.http import HttpResponse
from django.utils import timezone


class ExportListMixin:
    """
    Mixin genérico para exportar a PDF y Excel cualquier ListView.

    Respeta los filtros aplicados en ``get_queryset()`` (exporta TODOS los
    registros filtrados, no solo la página actual).

    Uso en la vista:
        class MiListView(ExportListMixin, ListView):
            export_title = 'Productos'
            export_fields = [
                ('Nombre', 'name'),                  # atributo simple
                ('Marca', 'brand.name'),             # atributo anidado
                ('Activo', lambda o: 'Sí' if o.is_active else 'No'),  # callable
            ]

    En la plantilla los botones apuntan a:  ?<filtros>&export=excel  /  ?<filtros>&export=pdf
    """
    export_fields = []          # lista de tuplas (encabezado, accessor)
    export_title = None         # título mostrado en el documento
    export_filename = None      # nombre base del archivo (sin extensión)

    # ------------------------------------------------------------------ #
    # Despacho
    # ------------------------------------------------------------------ #
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export')
        if fmt == 'excel':
            return self.export_excel()
        if fmt == 'pdf':
            return self.export_pdf()
        return super().get(request, *args, **kwargs)

    # ------------------------------------------------------------------ #
    # Utilidades
    # ------------------------------------------------------------------ #
    def get_export_title(self):
        return self.export_title or self.model._meta.verbose_name_plural.title()

    def get_export_filename(self):
        base = self.export_filename or self.get_export_title()
        stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
        return f'{base}_{stamp}'.replace(' ', '_')

    def get_export_headers(self):
        return [header for header, _ in self.export_fields]

    def resolve_value(self, obj, accessor):
        """Resuelve el valor de una celda a partir del accessor."""
        if callable(accessor):
            value = accessor(obj)
        else:
            value = obj
            for part in accessor.split('.'):
                value = getattr(value, part, '')
                if callable(value):
                    value = value()
        if value is None:
            return ''
        return value

    def get_export_rows(self):
        """Genera las filas (lista de listas de strings) a partir del queryset filtrado."""
        rows = []
        for obj in self.get_queryset():
            rows.append([str(self.resolve_value(obj, acc)) for _, acc in self.export_fields])
        return rows

    # ------------------------------------------------------------------ #
    # Excel
    # ------------------------------------------------------------------ #
    def export_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = self.get_export_title()[:31]

        headers = self.get_export_headers()
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='343A40')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for row in self.get_export_rows():
            ws.append(row)

        # Bordes + ancho automático
        widths = [len(h) for h in headers]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c, cell in enumerate(row):
                cell.border = border
                widths[c] = max(widths[c], len(str(cell.value or '')))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.xlsx"'
        wb.save(response)
        return response

    # ------------------------------------------------------------------ #
    # PDF
    # ------------------------------------------------------------------ #
    def export_pdf(self):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.pdf"'

        doc = SimpleDocTemplate(
            response, pagesize=landscape(A4),
            leftMargin=1.2 * cm, rightMargin=1.2 * cm,
            topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        )

        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)
        head_style = ParagraphStyle('cellHead', parent=styles['Normal'],
                                    fontSize=8, leading=10, textColor=colors.white, fontName='Helvetica-Bold')

        elements = [
            Paragraph(self.get_export_title(), styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal'],
            ),
            Spacer(1, 0.4 * cm),
        ]

        headers = [Paragraph(h, head_style) for h in self.get_export_headers()]
        data = [headers]
        for row in self.get_export_rows():
            data.append([Paragraph(value, cell_style) for value in row])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.4 * cm))
        elements.append(Paragraph(
            f"Total de registros: {len(data) - 1}", styles['Normal']
        ))

        doc.build(elements)
        return response
